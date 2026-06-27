from __future__ import annotations

import argparse
import html
import json
import re
import sys
import textwrap
import urllib.parse
from collections import Counter
from dataclasses import dataclass
from datetime import datetime
from functools import lru_cache
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from threading import Lock
from typing import Any

from docx import Document
from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent
REFLECTION_FILE = BASE_DIR / "성찰일지.xlsx"
COURSE_FILE = BASE_DIR / "2026-2 개설예정 교과목.xlsx"
CURRICULUM_FILE = BASE_DIR / "교육과정.docx"
ANALYTICS_FILE = Path("/tmp/analytics_events.jsonl") if Path("/tmp").exists() else BASE_DIR / "analytics_events.jsonl"
COURSE_REGISTRATION_URL = "https://portal.biocoss.ac.kr/"
IMAGE_FILES = {
    "/character.png": BASE_DIR / "new03.png",
    "/C01.png": BASE_DIR / "new03.png",
    "/C02.png": BASE_DIR / "new02.png",
    "/C03.png": BASE_DIR / "new01.png",
    "/C04.png": BASE_DIR / "new01.png",
    "/C_full06.png": BASE_DIR / "new03.png",
    "/C_bust01.png": BASE_DIR / "new01.png",
    "/C_bust02.png": BASE_DIR / "new02.png",
    "/C_bust03.png": BASE_DIR / "new01.png",
    "/new01.png": BASE_DIR / "new01.png",
    "/new02.png": BASE_DIR / "new02.png",
    "/new03.png": BASE_DIR / "new03.png",
}


REFLECTION_HEADER_ROW = 4
REFLECTION_QUESTION_ROW = 5
REFLECTION_DATA_START_ROW = 6
MAX_RECOMMENDED_COURSES = 6
MAX_RECOMMENDED_CURRICULA = 4
MAX_CTA_COURSES_PER_CURRICULUM = 3
ANALYTICS_LOCK = Lock()


STOPWORDS = {
    "그리고",
    "하지만",
    "통해",
    "대한",
    "관련",
    "중심",
    "분야",
    "생각",
    "학습",
    "수업",
    "진로",
    "목표",
    "활동",
    "과정",
    "계획",
    "교과목",
    "교육과정",
    "바이오헬스",
    "작성",
    "내용",
    "학생",
    "상담",
    "추천",
}


DOMAIN_KEYWORDS = {
    "디지털 헬스케어": ["디지털", "헬스케어", "플랫폼", "서비스", "데이터", "AI", "인공지능"],
    "의료 데이터 분석": ["데이터", "분석", "통계", "AI", "인공지능", "의료정보", "예측"],
    "재활 치료": ["재활", "치료", "운동", "고령", "건강관리", "돌봄"],
    "스마트 헬스 기기": ["스마트", "기기", "디바이스", "센서", "웨어러블", "의료기기"],
    "바이오헬스 디자인": ["디자인", "UX", "사용자", "서비스", "커뮤니케이션"],
    "바이오헬스 창업": ["창업", "사업", "비즈니스", "지식재산", "특허", "시장"],
}

SURVEY_INTEREST_AREAS = [
    "디지털 헬스케어",
    "의료 데이터 분석",
    "재활 치료",
    "스마트 헬스 기기",
    "바이오헬스 디자인",
    "바이오헬스 창업",
    "아직 탐색 중",
]
SURVEY_KEYWORDS = [
    "AI",
    "데이터 분석",
    "헬스케어",
    "재활",
    "운동",
    "돌봄",
    "웨어러블",
    "의료기기",
    "서비스 디자인",
    "창업",
    "비즈니스",
    "지역사회",
]
SURVEY_COMPETENCIES = ["문제해결", "데이터 활용", "콘텐츠 제작", "커뮤니케이션", "기획력", "연구/분석", "협업", "실무 프로젝트"]
SURVEY_CLASS_STYLES = ["실습 중심", "팀프로젝트", "이론 중심", "발표/토론", "포트폴리오 제작", "현장문제 해결"]
SURVEY_CONCERNS = ["진로 방향을 모르겠음", "전공을 어떻게 살릴지 고민됨", "취업 준비가 막막함", "포트폴리오가 부족함", "대학원/연구 관심", "창업 관심"]
SURVEY_INTENSITIES = ["부담 적은 탐색형", "역량 강화형", "진로 집중형", "도전형"]


@dataclass
class StudentRecord:
    student_id: str
    college: str
    affiliation: str
    track: str
    department: str
    answers: dict[str, str]
    question_texts: dict[str, str]


@dataclass
class Course:
    title: str
    code: str
    major: str
    track: str
    category: str
    level: str
    host: str
    format: str
    day: str
    time: str
    score: int = 0
    matched_terms: tuple[str, ...] = ()


@dataclass
class Curriculum:
    kind: str
    name: str
    overview: str
    composition: str
    score: int = 0
    matched_terms: tuple[str, ...] = ()


def normalize_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_student_id(value: Any) -> str:
    raw = normalize_cell(value)
    if raw.endswith(".0") and raw[:-2].isdigit():
        raw = raw[:-2]
    return re.sub(r"\s+", "", raw)


def mask_student_id(student_id: str) -> str:
    clean = normalize_student_id(student_id)
    if len(clean) <= 4:
        return "*" * len(clean)
    return f"{clean[:2]}{'*' * max(3, len(clean) - 5)}{clean[-3:]}"


def compact_text(value: str, limit: int = 280) -> str:
    text = re.sub(r"\s+", " ", value or "").strip()
    if len(text) <= limit:
        return text
    return text[: limit - 1].rstrip() + "…"


def split_sentences(text: str) -> list[str]:
    clean = re.sub(r"\s+", " ", text or "").strip()
    if not clean:
        return []
    marked = re.sub(r"(다\.|요\.|음\.|함\.|[.!?。])", r"\1\n", clean)
    parts = re.split(r"[\n\r]+", marked)
    return [part.strip(" .") for part in parts if part.strip(" .")]


def extract_terms(*texts: str) -> list[str]:
    joined = " ".join(texts)
    tokens = re.findall(r"[가-힣A-Za-z0-9+#·]{2,}", joined)
    weighted: list[str] = []
    for token in tokens:
        token = token.strip()
        if token in STOPWORDS:
            continue
        if len(token) < 2:
            continue
        weighted.append(token)

    for domain, terms in DOMAIN_KEYWORDS.items():
        if domain in joined or any(term in joined for term in terms):
            weighted.extend([domain, *terms])

    counts: dict[str, int] = {}
    for token in weighted:
        counts[token] = counts.get(token, 0) + 1
    return [
        term
        for term, _count in sorted(counts.items(), key=lambda item: (-item[1], item[0]))
        if term not in STOPWORDS
    ][:40]


@lru_cache(maxsize=1)
def load_student_records() -> tuple[dict[str, StudentRecord], list[str]]:
    wb = load_workbook(REFLECTION_FILE, read_only=True, data_only=True)
    ws = wb.active
    headers = [normalize_cell(cell.value) for cell in ws[REFLECTION_HEADER_ROW]]
    questions = [normalize_cell(cell.value) for cell in ws[REFLECTION_QUESTION_ROW]]
    idx = {header: pos for pos, header in enumerate(headers) if header}

    required = ["학번", "문항1", "문항2", "문항3", "문항4", "문항5"]
    missing = [name for name in required if name not in idx]
    if missing:
        raise RuntimeError(f"성찰일지 필수 컬럼을 찾을 수 없습니다: {', '.join(missing)}")

    records: dict[str, StudentRecord] = {}
    for row in ws.iter_rows(min_row=REFLECTION_DATA_START_ROW, values_only=True):
        student_id = normalize_student_id(row[idx["학번"]])
        if not student_id:
            continue
        answers = {
            key: normalize_cell(row[idx[key]])
            for key in ["문항1", "문항2", "문항3", "문항4", "문항5"]
        }
        question_texts = {
            key: questions[idx[key]] if idx[key] < len(questions) else key
            for key in ["문항1", "문항2", "문항3", "문항4", "문항5"]
        }
        records[student_id] = StudentRecord(
            student_id=student_id,
            college=normalize_cell(row[idx.get("대학", -1)]) if "대학" in idx else "",
            affiliation=normalize_cell(row[idx.get("소속", -1)]) if "소속" in idx else "",
            track=normalize_cell(row[idx.get("계열", -1)]) if "계열" in idx else "",
            department=normalize_cell(row[idx.get("학과", -1)]) if "학과" in idx else "",
            answers=answers,
            question_texts=question_texts,
        )
    return records, headers


@lru_cache(maxsize=1)
def load_courses() -> list[Course]:
    wb = load_workbook(COURSE_FILE, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers = [normalize_cell(value) for value in next(rows)]
    idx = {header: pos for pos, header in enumerate(headers) if header}

    def pick(row: tuple[Any, ...], name: str) -> str:
        return normalize_cell(row[idx[name]]) if name in idx and idx[name] < len(row) else ""

    courses: list[Course] = []
    for row in rows:
        title = pick(row, "교과목명")
        if not title:
            continue
        courses.append(
            Course(
                title=title,
                code=pick(row, "교과목코드"),
                major=pick(row, "성과용전공"),
                track=pick(row, "성과용트랙"),
                category=pick(row, "교과구분"),
                level=pick(row, "수준"),
                host=pick(row, "주관대학"),
                format=pick(row, "수업형태"),
                day=pick(row, "요일"),
                time=pick(row, "시간"),
            )
        )
    return courses


@lru_cache(maxsize=1)
def load_curricula() -> list[Curriculum]:
    doc = Document(CURRICULUM_FILE)
    curricula: list[Curriculum] = []
    for table in doc.tables:
        if not table.rows:
            continue
        headers = [cell.text.strip() for cell in table.rows[0].cells]
        header_idx = {name: pos for pos, name in enumerate(headers) if name}
        if "디그리명" not in header_idx:
            continue
        for row in table.rows[1:]:
            cells = [cell.text.strip() for cell in row.cells]

            def pick(name: str) -> str:
                pos = header_idx.get(name)
                return re.sub(r"\s+", " ", cells[pos]).strip() if pos is not None and pos < len(cells) else ""

            name = pick("디그리명")
            if not name:
                continue
            curricula.append(
                Curriculum(
                    kind=pick("유형"),
                    name=name,
                    overview=pick("개요"),
                    composition=pick("교육과정 구성"),
                )
            )
    return curricula


def score_text(text: str, terms: list[str], title_bonus: int = 1) -> tuple[int, tuple[str, ...]]:
    lower_text = text.lower()
    matched: list[str] = []
    score = 0
    for term in terms:
        lower_term = term.lower()
        if lower_term and lower_term in lower_text:
            matched.append(term)
            score += title_bonus if len(term) >= 4 else 1
    return score, tuple(dict.fromkeys(matched))


def recommend_courses(student: StudentRecord, terms: list[str]) -> list[Course]:
    profile_terms = extract_terms(student.department, student.track, student.affiliation)
    merged_terms = list(dict.fromkeys([*terms, *profile_terms]))
    scored: list[Course] = []
    for course in load_courses():
        title_score, title_matches = score_text(course.title, merged_terms, 4)
        meta_text = " ".join([course.major, course.track, course.category, course.level, course.host, course.format])
        meta_score, meta_matches = score_text(meta_text, merged_terms, 2)
        score = title_score + meta_score
        if "바이오헬스" in course.title:
            score += 1
        if student.department and student.department in meta_text:
            score += 3
        if student.track and student.track in meta_text:
            score += 2
        scored.append(
            Course(
                **{field: getattr(course, field) for field in course.__dataclass_fields__ if field not in {"score", "matched_terms"}},
                score=score,
                matched_terms=tuple(dict.fromkeys([*title_matches, *meta_matches])),
            )
        )
    scored.sort(key=lambda item: (-item.score, item.level != "초급", item.title))
    top = [course for course in scored if course.score > 0][:MAX_RECOMMENDED_COURSES]
    if len(top) < MAX_RECOMMENDED_COURSES:
        fallback = [course for course in scored if course not in top][: MAX_RECOMMENDED_COURSES - len(top)]
        top.extend(fallback)
    return top


def recommend_curricula(student: StudentRecord, terms: list[str]) -> list[Curriculum]:
    merged_terms = list(dict.fromkeys([*terms, *extract_terms(student.department, student.track, student.affiliation)]))
    scored: list[Curriculum] = []
    for curriculum in load_curricula():
        text = " ".join([curriculum.kind, curriculum.name, curriculum.overview, curriculum.composition])
        score, matches = score_text(text, merged_terms, 3)
        if "융합" in curriculum.name or "융합" in curriculum.kind:
            score += 1
        scored.append(
            Curriculum(
                kind=curriculum.kind,
                name=curriculum.name,
                overview=curriculum.overview,
                composition=curriculum.composition,
                score=score,
                matched_terms=matches,
            )
        )
    scored.sort(key=lambda item: (-item.score, item.name))
    top = [item for item in scored if item.score > 0][:MAX_RECOMMENDED_CURRICULA]
    if len(top) < MAX_RECOMMENDED_CURRICULA:
        top.extend([item for item in scored if item not in top][: MAX_RECOMMENDED_CURRICULA - len(top)])
    return top


def recommend_courses_for_curriculum(
    curriculum: Curriculum,
    courses: list[Course],
    student: StudentRecord,
    terms: list[str],
) -> list[Course]:
    curriculum_terms = extract_terms(curriculum.name, curriculum.overview, curriculum.composition)
    merged_terms = list(dict.fromkeys([*curriculum_terms, *terms, *extract_terms(student.department, student.track)]))
    scored: list[Course] = []
    for course in courses:
        course_text = " ".join([course.title, course.major, course.track, course.category, course.level, course.format])
        score, matches = score_text(course_text, merged_terms, 3)
        if course.title and course.title in curriculum.composition:
            score += 8
        if course.track and course.track in curriculum.name:
            score += 3
        if course.major and course.major in curriculum.name:
            score += 3
        if "바이오헬스" in course.title:
            score += 1
        if score <= 0:
            continue
        scored.append(
            Course(
                **{
                    field: getattr(course, field)
                    for field in course.__dataclass_fields__
                    if field not in {"score", "matched_terms"}
                },
                score=score,
                matched_terms=tuple(dict.fromkeys([*matches, *curriculum.matched_terms[:3]])),
            )
        )
    scored.sort(key=lambda item: (-item.score, item.level != "초급", item.title))
    return scored[:MAX_CTA_COURSES_PER_CURRICULUM]


def infer_focus_areas(student: StudentRecord, terms: list[str]) -> list[str]:
    joined = " ".join(student.answers.values())
    domains = [
        domain
        for domain, keywords in DOMAIN_KEYWORDS.items()
        if domain in joined or any(keyword in joined for keyword in keywords)
    ]
    if domains:
        return domains[:3]
    return [term for term in terms if len(term) >= 3][:3] or ["바이오헬스 진로 탐색"]


def summarize_goal(student: StudentRecord, terms: list[str]) -> str:
    focus = ", ".join(infer_focus_areas(student, terms))
    q5_sentences = split_sentences(student.answers.get("문항5", ""))
    q2_sentences = split_sentences(student.answers.get("문항2", ""))
    base_sentence = q5_sentences[0] if q5_sentences else ""
    interest_sentence = q2_sentences[0] if q2_sentences else ""
    profile = " / ".join(part for part in [student.affiliation, student.track, student.department] if part)

    if base_sentence and interest_sentence:
        return (
            f"{profile} 배경을 바탕으로 {focus}와 연결되는 진로를 탐색하고 있습니다. "
            f"성찰일지에서는 '{compact_text(interest_sentence, 90)}'라는 관심이 드러나며, "
            f"목표 설정에서는 '{compact_text(base_sentence, 120)}'라는 방향이 확인됩니다."
        )
    if base_sentence:
        return f"{profile} 배경을 바탕으로 {focus}와 연결되는 진로를 탐색하고 있으며, 목표 진술은 '{compact_text(base_sentence, 160)}'로 요약됩니다."
    return f"{profile} 배경과 성찰일지 응답을 기준으로 {focus} 관련 진로 목표를 더 구체화할 필요가 있습니다."


def build_report_for_student(student: StudentRecord) -> dict[str, Any]:
    all_answers = " ".join(student.answers.values())
    terms = extract_terms(all_answers, student.department, student.track, student.affiliation)
    courses = recommend_courses(student, terms)
    curricula = recommend_curricula(student, terms)
    all_courses = load_courses()
    curriculum_courses = {
        curriculum.name: recommend_courses_for_curriculum(curriculum, all_courses, student, terms)
        for curriculum in curricula
    }
    focus_areas = infer_focus_areas(student, terms)

    advice = [
        f"{focus_areas[0]}와 직접 연결되는 교과목을 먼저 수강해 관심 분야가 실제 학습 주제로 이어지는지 확인하세요.",
        "추천 교과목을 고를 때는 과목명만 보지 말고 성과용전공, 트랙, 수준, 수업형태를 함께 확인해 현재 준비도와 맞추는 것이 좋습니다.",
        "성찰일지에서 드러난 관심을 비교과 활동이나 프로젝트 주제로 바꿔, 다음 상담에서 포트폴리오형 증거로 가져올 수 있게 해보세요.",
    ]
    if len(focus_areas) > 1:
        advice.append(f"{focus_areas[0]}와 {focus_areas[1]} 중 어느 쪽이 더 오래 탐색할 주제인지 작은 과제나 인터뷰로 비교해보세요.")

    reflection_questions = [
        f"추천 교과목 중 가장 먼저 수강하고 싶은 과목은 무엇이며, 그 과목이 {focus_areas[0]} 탐색에 어떤 도움을 줄까요?",
        "성찰일지에서 말한 관심사가 실제 직무나 연구 주제로 이어지려면 어떤 경험이 더 필요할까요?",
        "추천 교육과정 중 본인의 학과/계열과 가장 자연스럽게 연결되는 과정은 무엇이고, 연결이 약한 과정은 왜 그런가요?",
        "다음 학기 말에 '진로 탐색이 진전됐다'고 판단할 수 있는 증거 1개는 무엇인가요?",
        "상담자와 다시 확인해야 할 불확실한 점이나 부담되는 조건은 무엇인가요?",
    ]

    return {
        "student": student,
        "masked_student_id": mask_student_id(student.student_id),
        "focus_areas": focus_areas,
        "terms": terms[:12],
        "goal_summary": summarize_goal(student, terms),
        "courses": courses,
        "curricula": curricula,
        "curriculum_courses": curriculum_courses,
        "advice": advice,
        "reflection_questions": reflection_questions,
    }


def build_report(student_id: str) -> dict[str, Any]:
    records, _headers = load_student_records()
    normalized_id = normalize_student_id(student_id)
    student = records.get(normalized_id)
    if not student:
        sample_ids = list(records.keys())[:5]
        raise KeyError(
            "해당 학번을 찾을 수 없습니다. 입력값의 공백을 확인하거나 성찰일지.xlsx에 있는 익명화 학번을 사용해 주세요. "
            f"데이터에는 {len(records)}명의 기록이 있습니다."
        )

    return build_report_for_student(student)


def first_param(params: dict[str, list[str]], key: str) -> str:
    return params.get(key, [""])[0].strip()


def build_survey_student(params: dict[str, list[str]]) -> StudentRecord:
    interest_area = first_param(params, "interest_area")
    keywords = [item.strip() for item in params.get("keywords", []) if item.strip()]
    competency = first_param(params, "competency")
    class_style = first_param(params, "class_style")
    concern = first_param(params, "concern")
    intensity = first_param(params, "intensity")
    free_text = first_param(params, "free_text")
    department = first_param(params, "department") or "미정"
    student_id = normalize_student_id(first_param(params, "student_id")) or f"survey-{datetime.now().strftime('%Y%m%d%H%M%S')}"

    if not interest_area or not keywords or not competency or not class_style or not concern:
        raise ValueError("관심 분야, 관심 키워드, 키우고 싶은 역량, 선호 수업 방식, 현재 진로 고민을 모두 입력해 주세요.")

    keyword_text = ", ".join(keywords)
    answers = {
        "문항1": f"관심 분야는 {interest_area}이며, 현재 {department} 배경에서 진로를 탐색하고 있습니다.",
        "문항2": f"흥미 있는 키워드는 {keyword_text}입니다. {free_text}",
        "문항3": f"이번 학기에 키우고 싶은 역량은 {competency}입니다.",
        "문항4": f"선호하는 수업 방식은 {class_style}이며, 추천 강도는 {intensity or '역량 강화형'}을 원합니다.",
        "문항5": f"현재 진로 고민은 {concern}입니다. {free_text}",
    }
    question_texts = {
        "문항1": "관심 있는 진로/분야",
        "문항2": "흥미 있는 키워드",
        "문항3": "키우고 싶은 역량",
        "문항4": "선호하는 수업 방식",
        "문항5": "현재 진로 고민",
    }
    return StudentRecord(
        student_id=student_id,
        college="문항 응답",
        affiliation="간단 응답 기반",
        track=interest_area,
        department=department,
        answers=answers,
        question_texts=question_texts,
    )


def build_survey_report(params: dict[str, list[str]]) -> dict[str, Any]:
    report = build_report_for_student(build_survey_student(params))
    report["survey_based"] = True
    return report


def course_reason(course: Course, student: StudentRecord) -> str:
    matches = ", ".join(course.matched_terms[:5]) if course.matched_terms else "바이오헬스 기초/확장 학습"
    meta = " / ".join(part for part in [course.major, course.track, course.level, course.format] if part)
    return f"{matches} 키워드와 연결됩니다. {meta} 정보를 상담자가 확인해 학생의 현재 준비도와 맞출 수 있습니다."


def curriculum_reason(curriculum: Curriculum) -> str:
    matches = ", ".join(curriculum.matched_terms[:5]) if curriculum.matched_terms else "융합형 바이오헬스 학습"
    return f"{matches} 키워드와 연결되며, {compact_text(curriculum.overview, 120)}"


def render_markdown(report: dict[str, Any]) -> str:
    student: StudentRecord = report["student"]
    courses: list[Course] = report["courses"]
    curricula: list[Curriculum] = report["curricula"]
    curriculum_courses: dict[str, list[Course]] = report.get("curriculum_courses", {})

    lines = [
        "# 진로상담 리포트 초안",
        "",
        "## 학생 기본 정보",
        f"- 학번: {report['masked_student_id']} (공개용 마스킹)",
        f"- 소속: {student.affiliation or '미정'}",
        f"- 계열: {student.track or '미정'}",
        f"- 학과: {student.department or '미정'}",
        "",
        "## 진로목표 요약",
        report["goal_summary"],
        "",
        "## 성찰일지 핵심 참고",
    ]
    for key in ["문항1", "문항2", "문항3", "문항4", "문항5"]:
        lines.append(f"- {key}: {compact_text(student.answers.get(key, ''), 180) or '응답 없음'}")

    lines.extend(["", "## 추천 교과목"])
    for index, course in enumerate(courses, 1):
        schedule = " ".join(part for part in [course.day, course.time] if part)
        lines.extend(
            [
                f"{index}. **{course.title}**",
                f"   - 코드/수준: {course.code or '미정'} / {course.level or '미정'}",
                f"   - 전공/트랙: {course.major or '미정'} / {course.track or '미정'}",
                f"   - 수업: {course.category or '미정'}, {course.format or '미정'}, {schedule or '시간 미정'}",
                f"   - 추천 근거: {course_reason(course, student)}",
            ]
        )

    lines.extend(["", "## 추천 교육과정"])
    for index, curriculum in enumerate(curricula, 1):
        cta_courses = curriculum_courses.get(curriculum.name, [])
        lines.extend(
            [
                f"{index}. **{curriculum.name}** ({curriculum.kind or '유형 미정'})",
                f"   - 개요: {compact_text(curriculum.overview, 220)}",
                f"   - 추천 근거: {curriculum_reason(curriculum)}",
                "   - 바로 확인할 2026-2 개설 교과목:",
            ]
        )
        if cta_courses:
            for course in cta_courses:
                schedule = " ".join(part for part in [course.day, course.time] if part)
                lines.append(
                    f"     - {course.title} ({course.level or '수준 미정'}, {course.format or '수업형태 미정'}, {schedule or '시간 미정'})"
                )
        else:
            lines.append("     - 연결 가능한 개설 교과목을 상담자가 추가 확인해야 합니다.")

    lines.extend(["", "## 진로 탐색 조언"])
    lines.extend(f"- {item}" for item in report["advice"])

    lines.extend(["", "## 추가 성찰 질문"])
    lines.extend(f"{index}. {question}" for index, question in enumerate(report["reflection_questions"], 1))

    lines.extend(
        [
            "",
            "## 상담자 검토 체크리스트",
            "- [ ] 추천 교과목이 2026-2 개설예정 교과목 파일에 실제로 있는지 확인",
            "- [ ] 추천 교육과정이 교육과정 문서의 설명과 맞는지 확인",
            "- [ ] 학생에게 과도하게 단정적인 표현이 없는지 확인",
            "- [ ] 공개 또는 공유본에서 학번 등 식별 정보가 마스킹되어 있는지 확인",
        ]
    )
    return "\n".join(lines) + "\n"


def analytics_profile(student: StudentRecord) -> dict[str, str]:
    return {
        "masked_student_id": mask_student_id(student.student_id),
        "college": student.college or "미정",
        "affiliation": student.affiliation or "미정",
        "track": student.track or "미정",
        "department": student.department or "미정",
    }


def write_analytics_event(event_type: str, student: StudentRecord | None = None, extra: dict[str, Any] | None = None) -> None:
    event: dict[str, Any] = {
        "type": event_type,
        "timestamp": datetime.now().isoformat(timespec="seconds"),
    }
    if student:
        event.update(analytics_profile(student))
    if extra:
        event.update(extra)
    with ANALYTICS_LOCK:
        with ANALYTICS_FILE.open("a", encoding="utf-8") as file:
            file.write(json.dumps(event, ensure_ascii=False) + "\n")


def read_analytics_events() -> list[dict[str, Any]]:
    if not ANALYTICS_FILE.exists():
        return []
    events: list[dict[str, Any]] = []
    with ANALYTICS_FILE.open("r", encoding="utf-8") as file:
        for line in file:
            line = line.strip()
            if not line:
                continue
            try:
                events.append(json.loads(line))
            except json.JSONDecodeError:
                continue
    return events


def counter_for(events: list[dict[str, Any]], field: str) -> Counter[str]:
    return Counter(str(event.get(field) or "미정") for event in events)


def render_count_rows(counter: Counter[str], total: int) -> str:
    if not counter:
        return "<tr><td>데이터 없음</td><td>0</td><td><div class='bar'><span style='width:0%'></span></div></td></tr>"
    rows = []
    for label, count in counter.most_common(8):
        percent = round((count / total) * 100) if total else 0
        rows.append(
            f"""
            <tr>
              <td>{html.escape(label)}</td>
              <td>{count}</td>
              <td><div class="bar"><span style="width:{percent}%"></span></div></td>
            </tr>
            """
        )
    return "".join(rows)


def render_recent_events(events: list[dict[str, Any]]) -> str:
    if not events:
        return "<tr><td colspan='6'>아직 기록된 활동이 없습니다.</td></tr>"
    rows = []
    labels = {"lookup": "리포트 조회", "course_click": "수강신청 클릭"}
    for event in reversed(events[-20:]):
        rows.append(
            f"""
            <tr>
              <td>{html.escape(event.get('timestamp', ''))}</td>
              <td>{html.escape(labels.get(event.get('type'), event.get('type', '기타')))}</td>
              <td>{html.escape(event.get('masked_student_id', '미정'))}</td>
              <td>{html.escape(event.get('college', '미정'))}</td>
              <td>{html.escape(event.get('track', '미정'))}</td>
              <td>{html.escape(event.get('department', '미정'))}</td>
            </tr>
            """
        )
    return "".join(rows)


def event_date(event: dict[str, Any]) -> str:
    timestamp = str(event.get("timestamp") or "")
    if len(timestamp) >= 10:
        return timestamp[:10]
    return "날짜 미정"


def render_daily_rows(events: list[dict[str, Any]]) -> str:
    if not events:
        return "<tr><td colspan='5'>아직 기록된 날짜별 활동이 없습니다.</td></tr>"
    daily: dict[str, dict[str, Any]] = {}
    for event in events:
        day = event_date(event)
        bucket = daily.setdefault(day, {"lookup": 0, "course_click": 0, "users": set()})
        event_type = event.get("type")
        if event_type in ("lookup", "course_click"):
            bucket[event_type] += 1
        masked_id = event.get("masked_student_id")
        if masked_id and masked_id != "미정":
            bucket["users"].add(masked_id)
    rows = []
    for day in sorted(daily.keys(), reverse=True)[:14]:
        bucket = daily[day]
        lookup_count = bucket["lookup"]
        click_count = bucket["course_click"]
        conversion = round((click_count / lookup_count) * 100, 1) if lookup_count else 0
        rows.append(
            f"""
            <tr>
              <td>{html.escape(day)}</td>
              <td>{lookup_count}</td>
              <td>{len(bucket['users'])}</td>
              <td>{click_count}</td>
              <td>{conversion}%</td>
            </tr>
            """
        )
    return "".join(rows)


def render_admin_html() -> str:
    events = read_analytics_events()
    lookups = [event for event in events if event.get("type") == "lookup"]
    clicks = [event for event in events if event.get("type") == "course_click"]
    unique_users = len({event.get("masked_student_id") for event in lookups if event.get("masked_student_id")})
    conversion = round((len(clicks) / len(lookups)) * 100, 1) if lookups else 0
    latest = events[-1]["timestamp"] if events else "기록 없음"
    college_rows = render_count_rows(counter_for(lookups, "college"), len(lookups))
    affiliation_rows = render_count_rows(counter_for(lookups, "affiliation"), len(lookups))
    track_rows = render_count_rows(counter_for(lookups, "track"), len(lookups))
    department_rows = render_count_rows(counter_for(lookups, "department"), len(lookups))
    daily_rows = render_daily_rows(events)
    recent_rows = render_recent_events(events)
    return f"""<!doctype html>
<html lang="ko">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>관리자 대시보드 | FuturesImpactLAB</title>
  <style>
    :root {{ --ink:#191a23; --muted:#5c5f69; --line:#191a23; --panel:#f3f3f3; --accent:#b9ff66; --shadow:0 5px 0 #191a23; }}
    * {{ box-sizing:border-box; }}
    body {{ margin:0; font-family:Arial, "Malgun Gothic", sans-serif; color:var(--ink); background:#fbfbfb; line-height:1.5; }}
    header, main {{ max-width:1180px; margin:0 auto; padding:24px; }}
    .top {{ display:flex; align-items:flex-end; justify-content:space-between; gap:16px; padding:24px; border:2px solid var(--line); border-radius:8px; background:var(--panel); box-shadow:var(--shadow); }}
    h1 {{ margin:0 0 8px; font-size:34px; letter-spacing:0; }}
    h2 {{ margin:0 0 14px; font-size:20px; letter-spacing:0; }}
    p {{ margin:0; color:#343640; }}
    a {{ color:var(--ink); font-weight:800; }}
    .home-link {{ display:inline-flex; align-items:center; min-height:44px; padding:0 16px; border:2px solid var(--line); border-radius:8px; background:var(--accent); text-decoration:none; box-shadow:0 3px 0 #191a23; white-space:nowrap; }}
    .stats {{ display:grid; grid-template-columns:repeat(4, minmax(0, 1fr)); gap:14px; margin:22px 0; }}
    .stat, section {{ border:2px solid var(--line); border-radius:8px; background:#fff; box-shadow:var(--shadow); }}
    .stat {{ padding:18px; min-height:112px; }}
    .label {{ display:block; margin-bottom:8px; color:var(--muted); font-size:13px; font-weight:800; }}
    .value {{ font-size:30px; font-weight:900; }}
    .dashboard {{ display:grid; grid-template-columns:repeat(2, minmax(0, 1fr)); gap:18px; }}
    section {{ padding:20px; overflow:auto; }}
    table {{ width:100%; border-collapse:collapse; font-size:14px; }}
    th, td {{ padding:10px 8px; border-bottom:1px solid #dedede; text-align:left; vertical-align:middle; }}
    th {{ font-size:12px; color:var(--muted); }}
    .bar {{ width:100%; height:12px; border:1.5px solid var(--line); border-radius:999px; background:#fff; overflow:hidden; }}
    .bar span {{ display:block; height:100%; background:var(--accent); }}
    .wide {{ grid-column:1 / -1; }}
    .ideas {{ display:grid; grid-template-columns:repeat(3, minmax(0, 1fr)); gap:12px; }}
    .idea {{ padding:14px; border:1.5px solid var(--line); border-radius:8px; background:var(--panel); }}
    .idea strong {{ display:block; margin-bottom:4px; }}
    @media (max-width: 820px) {{ .top, .stats, .dashboard, .ideas {{ grid-template-columns:1fr; display:grid; }} .top {{ align-items:stretch; }} .home-link {{ justify-content:center; }} }}
  </style>
</head>
<body>
  <header>
    <div class="top">
      <div>
        <span class="label">Admin</span>
        <h1>사용자 행동 대시보드</h1>
        <p>학번으로 리포트를 조회한 사용자와 실제 수강신청 페이지로 이동한 사용자를 확인합니다.</p>
      </div>
      <a class="home-link" href="/">리포트 화면으로</a>
    </div>
  </header>
  <main>
    <div class="stats">
      <div class="stat"><span class="label">리포트 조회</span><div class="value">{len(lookups)}</div></div>
      <div class="stat"><span class="label">고유 조회 사용자</span><div class="value">{unique_users}</div></div>
      <div class="stat"><span class="label">수강신청 클릭</span><div class="value">{len(clicks)}</div></div>
      <div class="stat"><span class="label">클릭 전환율</span><div class="value">{conversion}%</div></div>
    </div>
    <div class="dashboard">
      <section><h2>대학별 조회</h2><table><thead><tr><th>대학</th><th>건수</th><th>비중</th></tr></thead><tbody>{college_rows}</tbody></table></section>
      <section><h2>소속별 조회</h2><table><thead><tr><th>소속</th><th>건수</th><th>비중</th></tr></thead><tbody>{affiliation_rows}</tbody></table></section>
      <section><h2>계열별 조회</h2><table><thead><tr><th>계열</th><th>건수</th><th>비중</th></tr></thead><tbody>{track_rows}</tbody></table></section>
      <section><h2>학과별 조회</h2><table><thead><tr><th>학과</th><th>건수</th><th>비중</th></tr></thead><tbody>{department_rows}</tbody></table></section>
      <section class="wide"><h2>날짜별 현황</h2><table><thead><tr><th>날짜</th><th>리포트 조회</th><th>고유 사용자</th><th>수강신청 클릭</th><th>클릭 전환율</th></tr></thead><tbody>{daily_rows}</tbody></table></section>
      <section class="wide"><h2>최근 활동</h2><table><thead><tr><th>시간</th><th>활동</th><th>학번</th><th>대학</th><th>계열</th><th>학과</th></tr></thead><tbody>{recent_rows}</tbody></table></section>
      <section class="wide"><h2>추가로 보면 좋은 분석</h2><div class="ideas">
        <div class="idea"><strong>계열별 클릭 전환율</strong><p>조회는 많지만 클릭이 낮은 계열을 찾아 안내 문구나 추천 과목 구성을 조정할 수 있습니다.</p></div>
        <div class="idea"><strong>학과별 인기 추천 영역</strong><p>리포트의 관심 키워드를 함께 저장하면 학과별로 어떤 진로 주제가 많이 나오는지 볼 수 있습니다.</p></div>
        <div class="idea"><strong>상담 후속 대상</strong><p>조회 후 수강신청 클릭이 없는 사용자를 묶어 추가 상담이나 안내 메시지 대상으로 볼 수 있습니다.</p></div>
      </div></section>
    </div>
    <p style="margin-top:18px;color:#5c5f69;font-size:13px;">최근 업데이트: {html.escape(latest)} · 기록 파일: {html.escape(ANALYTICS_FILE.name)}</p>
  </main>
</body>
</html>"""


def render_select(name: str, label: str, options: list[str], required: bool = True) -> str:
    option_tags = ['<option value="">선택해 주세요</option>']
    option_tags.extend(f'<option value="{html.escape(option)}">{html.escape(option)}</option>' for option in options)
    required_attr = " required" if required else ""
    return f"""
      <label>
        <span>{html.escape(label)}</span>
        <select name="{html.escape(name)}"{required_attr}>{''.join(option_tags)}</select>
      </label>
    """


def render_checkbox_group(name: str, label: str, options: list[str]) -> str:
    items = "".join(
        f"""
        <label class="choice">
          <input type="checkbox" name="{html.escape(name)}" value="{html.escape(option)}">
          <span>{html.escape(option)}</span>
        </label>
        """
        for option in options
    )
    return f"""
      <fieldset>
        <legend>{html.escape(label)}</legend>
        <div class="choice-grid">{items}</div>
      </fieldset>
    """


def render_survey_form(student_id: str = "") -> str:
    return f"""
    <section class="survey-panel">
      <div>
        <span class="section-kicker">Quick Survey</span>
        <h2>성찰일지 데이터가 없을 때 간단 응답으로 추천 받기</h2>
        <p>아래 문항에 답하면 교과목/교육과정 추천에 필요한 최소 정보를 바탕으로 상담용 리포트를 바로 생성합니다.</p>
      </div>
      <form class="survey-form" method="post">
        <input type="hidden" name="mode" value="survey">
        <label>
          <span>학번 또는 구분값</span>
          <input name="student_id" value="{html.escape(student_id)}" placeholder="선택 입력">
        </label>
        <label>
          <span>학과/전공</span>
          <input name="department" placeholder="예: 간호학과, 작업치료학과, 미정">
        </label>
        {render_select("interest_area", "관심 있는 진로/분야", SURVEY_INTEREST_AREAS)}
        {render_checkbox_group("keywords", "흥미 있는 키워드 2~3개", SURVEY_KEYWORDS)}
        {render_select("competency", "이번 학기에 키우고 싶은 역량", SURVEY_COMPETENCIES)}
        {render_select("class_style", "선호하는 수업 방식", SURVEY_CLASS_STYLES)}
        {render_select("concern", "현재 가장 고민되는 것", SURVEY_CONCERNS)}
        {render_select("intensity", "희망하는 추천 강도", SURVEY_INTENSITIES, required=False)}
        <label class="wide-field">
          <span>자유 응답 한 줄</span>
          <input name="free_text" placeholder="예: 데이터 기반 건강관리 서비스에 관심이 있어요.">
        </label>
        <button type="submit">간단 응답으로 리포트 생성</button>
      </form>
    </section>
    """


def render_html(report: dict[str, Any] | None = None, error: str = "", student_id: str = "", survey: bool = False) -> str:
    body = ""
    if error:
        body = f"<section class='notice error'>{html.escape(error)}</section>"
        if survey:
            body += render_survey_form(student_id)
    elif report:
        student: StudentRecord = report["student"]
        courses: list[Course] = report["courses"]
        curricula: list[Curriculum] = report["curricula"]
        curriculum_courses: dict[str, list[Course]] = report.get("curriculum_courses", {})
        course_cards = "".join(
            f"""
            <article class="item">
              <div class="item-title">{html.escape(course.title)}</div>
              <div class="meta">{html.escape(' / '.join(part for part in [course.major, course.track, course.level, course.format] if part))}</div>
              <p>{html.escape(course_reason(course, student))}</p>
            </article>
            """
            for course in courses
        )
        curriculum_card_parts = []
        for curriculum in curricula:
            cta_courses = curriculum_courses.get(curriculum.name, [])
            cta_items = "".join(
                f"""
                <li>
                  <strong>{html.escape(course.title)}</strong>
                  <span>{html.escape(' / '.join(part for part in [course.level, course.format, ' '.join(p for p in [course.day, course.time] if p)] if part) or '수업 정보 미정')}</span>
                </li>
                """
                for course in cta_courses
            )
            if not cta_items:
                cta_items = "<li><strong>추가 확인 필요</strong><span>교육과정과 연결되는 개설 교과목을 상담자가 확인합니다.</span></li>"
            curriculum_card_parts.append(
                f"""
            <article class="item">
              <div class="item-title">{html.escape(curriculum.name)}</div>
              <div class="meta">{html.escape(curriculum.kind or '유형 미정')}</div>
              <p>{html.escape(compact_text(curriculum.overview, 180))}</p>
              <p class="reason">{html.escape(curriculum_reason(curriculum))}</p>
              <div class="cta-title">바로 확인할 2026-2 개설 교과목</div>
              <ul class="cta-list">{cta_items}</ul>
            </article>
            """
            )
        curriculum_cards = "".join(curriculum_card_parts)
        advice = "".join(f"<li>{html.escape(item)}</li>" for item in report["advice"])
        questions = "".join(f"<li>{html.escape(item)}</li>" for item in report["reflection_questions"])
        body = f"""
        <section class="summary">
          <div>
            <span class="label">공개용 학번</span>
            <strong>{html.escape(report['masked_student_id'])}</strong>
          </div>
          <div>
            <span class="label">소속/계열/학과</span>
            <strong>{html.escape(' / '.join(part for part in [student.affiliation, student.track, student.department] if part) or '미정')}</strong>
          </div>
        </section>
        {('<section class="notice"><strong>간단 응답 기반 리포트입니다.</strong><p>성찰일지 원자료가 없는 상황에서 입력한 문항 응답을 바탕으로 생성했습니다. 상담 시 실제 이수 상황과 관심 변화를 함께 확인해 주세요.</p></section>' if report.get('survey_based') else '')}
        <section>
          <h2>진로목표 요약</h2>
          <p>{html.escape(report['goal_summary'])}</p>
        </section>
        <section id="recommended-courses">
          <div class="section-head">
            <div>
              <span class="section-kicker">Courses</span>
              <h2>추천 교과목</h2>
            </div>
            <img class="section-mascot" src="/new02.png" alt="">
          </div>
          <div class="grid">{course_cards}</div>
        </section>
        <section>
          <div class="section-head">
            <div>
              <span class="section-kicker">Curriculum</span>
              <h2>추천 교육과정</h2>
            </div>
            <img class="section-mascot" src="/new01.png" alt="">
          </div>
          <div class="grid">{curriculum_cards}</div>
        </section>
        <section>
          <div class="guidance-panel">
            <div>
              <div class="section-head compact">
                <div>
                  <span class="section-kicker">Reflection</span>
                  <h2>진로 탐색 조언</h2>
                </div>
              </div>
              <ul>{advice}</ul>
              <h2 class="subhead">추가 성찰 질문</h2>
              <ol>{questions}</ol>
            </div>
            <img class="guidance-mascot" src="/new01.png" alt="">
          </div>
        </section>
        <section class="check-panel">
          <div>
            <span class="section-kicker">Check</span>
            <h2>꼭 확인해야 할 점</h2>
          </div>
          <p>이 리포트는 최종 상담 결과가 아니라, 상담자와 학생이 함께 검토하기 위한 초안입니다. 추천 교과목은 실제 개설 여부와 수강 가능 여부를 기준으로 제안되었지만, 선수과목, 시간표, 졸업요건, 개인별 이수 상황은 반드시 추가로 확인해야 합니다.</p>
          <p>AI가 제안한 방향은 출발점입니다. 리포트를 그대로 받아들이기보다는 자신의 관심과 상황에 맞게 상담자와 함께 수정하고 조정해 주세요.</p>
        </section>
        <section class="final-cta">
          <div>
            <span class="label">다음 행동</span>
            <h2>추천 교과목을 확인했다면, 이제 수강신청하러 가볼까요?</h2>
            <p>관심 있는 과목을 표시해보고, 상담자와 함께 우선순위를 정한 뒤 다음 학기 수강계획으로 이어가면 됩니다.</p>
          </div>
          <a class="cta-button" href="/go-course?student_id={urllib.parse.quote(student.student_id)}">수강신청 바로가기</a>
        </section>
        """
    elif survey:
        body = render_survey_form(student_id)

    return f"""<!doctype html>
<html lang="ko">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>FuturesImpactLAB 진로상담 리포트</title>
  <style>
    :root {{
      --ink: #191a23;
      --muted: #5c5f69;
      --line: #191a23;
      --panel: #f3f3f3;
      --accent: #b9ff66;
      --accent-2: #191a23;
      --danger: #9f2d2d;
      --shadow: 0 6px 0 #191a23;
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      font-family: Arial, "Malgun Gothic", sans-serif;
      color: var(--ink);
      background: #fbfbfb;
      line-height: 1.55;
    }}
    header {{
      max-width: 1180px;
      margin: 0 auto;
      padding: 28px 24px 12px;
      background: transparent;
    }}
    .hero {{
      display: grid;
      grid-template-columns: minmax(0, 1.05fr) minmax(260px, 0.95fr);
      gap: 24px;
      align-items: center;
      padding: 34px;
      border: 2px solid var(--line);
      border-radius: 8px;
      background: #f3f3f3;
      box-shadow: var(--shadow);
    }}
    .eyebrow {{
      display: inline-flex;
      align-items: center;
      width: fit-content;
      margin-bottom: 14px;
      padding: 6px 12px;
      border: 1.5px solid var(--line);
      border-radius: 999px;
      background: var(--accent);
      font-size: 13px;
      font-weight: 700;
    }}
    h1 {{ margin: 0 0 14px; font-size: 44px; line-height: 1.08; letter-spacing: 0; }}
    h2 {{ margin: 0 0 12px; font-size: 22px; letter-spacing: 0; }}
    p {{ margin: 0; }}
    .hero p {{ max-width: 620px; color: #343640; font-size: 17px; }}
    .intro-panel {{
      display: grid;
      grid-template-columns: minmax(0, 1.2fr) minmax(260px, 0.8fr);
      gap: 18px;
      max-width: 1180px;
      margin: 18px auto 0;
      padding: 0 24px;
    }}
    .intro-card {{
      border: 2px solid var(--line);
      border-radius: 8px;
      padding: 22px;
      background: #ffffff;
      box-shadow: 0 4px 0 #191a23;
    }}
    .intro-card.accent {{
      background: var(--accent);
    }}
    .intro-card p + p {{
      margin-top: 12px;
    }}
    .intro-list {{
      display: grid;
      gap: 10px;
      margin: 0;
      padding: 0;
      list-style: none;
    }}
    .intro-list li {{
      padding: 10px 0;
      border-bottom: 1.5px solid rgba(25, 26, 35, 0.22);
      font-weight: 700;
    }}
    .intro-list li:last-child {{
      border-bottom: 0;
    }}
    .hero-art {{
      display: flex;
      justify-content: center;
      align-items: flex-end;
      min-height: 330px;
      border: 2px solid var(--line);
      border-radius: 8px;
      background: #061923;
      overflow: hidden;
      box-shadow: 0 4px 0 #191a23;
    }}
    .hero-art img {{
      width: 100%;
      height: 330px;
      object-fit: cover;
      object-position: center 34%;
      padding: 0;
    }}
    main {{ max-width: 1180px; margin: 0 auto; padding: 20px 24px 42px; }}
    form {{
      display: grid;
      grid-template-columns: minmax(220px, 1fr) auto;
      gap: 12px;
      padding: 18px;
      border: 2px solid var(--line);
      border-radius: 8px;
      background: #ffffff;
      box-shadow: 0 4px 0 #191a23;
      margin-bottom: 24px;
    }}
    input,
    select {{
      min-height: 52px;
      border: 2px solid var(--line);
      border-radius: 8px;
      padding: 0 16px;
      font-size: 16px;
      background: #fff;
    }}
    button {{
      min-height: 52px;
      border: 2px solid var(--line);
      border-radius: 8px;
      padding: 0 24px;
      font-size: 16px;
      font-weight: 700;
      color: #fff;
      background: var(--ink);
      cursor: pointer;
    }}
    button:hover {{ background: #2f3140; }}
    .survey-link {{
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 16px;
      margin: -8px 0 22px;
      padding: 18px 20px;
      border: 2px solid var(--line);
      border-radius: 8px;
      background: #fff;
      box-shadow: 0 4px 0 #191a23;
    }}
    .survey-link p {{ color: #343640; }}
    .survey-panel {{
      display: grid;
      gap: 18px;
      border: 2px solid var(--line);
      border-radius: 8px;
      padding: 24px;
      background: #f3f3f3;
      box-shadow: var(--shadow);
    }}
    .survey-panel > div > p {{
      max-width: 820px;
      color: #343640;
    }}
    .survey-form {{
      grid-template-columns: repeat(2, minmax(0, 1fr));
      margin: 0;
      box-shadow: none;
    }}
    .survey-form label,
    .survey-form fieldset {{
      display: grid;
      gap: 8px;
      margin: 0;
      min-width: 0;
      border: 0;
      padding: 0;
      font-weight: 700;
    }}
    .survey-form fieldset,
    .survey-form .wide-field,
    .survey-form button {{
      grid-column: 1 / -1;
    }}
    .survey-form legend {{
      margin-bottom: 8px;
      font-weight: 700;
    }}
    .choice-grid {{
      display: grid;
      grid-template-columns: repeat(4, minmax(0, 1fr));
      gap: 8px;
    }}
    .choice {{
      display: flex !important;
      align-items: center;
      gap: 8px !important;
      min-height: 44px;
      padding: 8px 10px !important;
      border: 2px solid var(--line) !important;
      border-radius: 8px;
      background: #fff;
      font-weight: 700;
    }}
    .choice input {{
      min-height: auto;
      width: 18px;
      height: 18px;
      padding: 0;
    }}
    section {{
      border-top: 0;
      padding: 22px 0;
    }}
    .summary {{
      display: grid;
      grid-template-columns: repeat(2, minmax(0, 1fr));
      gap: 16px;
      border-top: 0;
      padding-top: 4px;
    }}
    .summary > div {{
      background: var(--accent);
      border: 2px solid var(--line);
      border-radius: 8px;
      padding: 18px;
      box-shadow: 0 4px 0 #191a23;
    }}
    .label {{
      display: block;
      color: #343640;
      font-size: 13px;
      font-weight: 700;
      margin-bottom: 4px;
    }}
    .grid {{
      display: grid;
      grid-template-columns: repeat(2, minmax(0, 1fr));
      gap: 18px;
    }}
    .section-head {{
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 18px;
      min-height: 104px;
      margin-bottom: 18px;
      padding: 18px 22px;
      border: 2px solid var(--line);
      border-radius: 8px;
      background: #ffffff;
      box-shadow: 0 4px 0 #191a23;
      overflow: hidden;
    }}
    .section-head.compact {{
      min-height: auto;
      margin-bottom: 10px;
      padding: 0;
      border: 0;
      box-shadow: none;
      background: transparent;
    }}
    .section-kicker {{
      display: inline-flex;
      width: fit-content;
      margin-bottom: 8px;
      padding: 5px 10px;
      border: 1.5px solid var(--line);
      border-radius: 8px;
      background: var(--accent);
      font-size: 13px;
      font-weight: 800;
    }}
    .section-mascot {{
      width: 168px;
      height: 104px;
      object-fit: cover;
      object-position: center center;
      align-self: center;
      border: 2px solid var(--line);
      border-radius: 8px;
      background: #061923;
      box-shadow: 0 3px 0 #191a23;
    }}
    .item {{
      border: 2px solid var(--line);
      border-radius: 8px;
      padding: 20px;
      background: #fff;
      box-shadow: var(--shadow);
    }}
    .item:nth-child(even) {{ background: #f3f3f3; }}
    .item-title {{ font-weight: 800; margin-bottom: 6px; font-size: 18px; }}
    .meta {{
      display: inline-block;
      color: var(--ink);
      background: var(--accent);
      border-radius: 8px;
      padding: 3px 8px;
      font-size: 13px;
      font-weight: 700;
      margin-bottom: 10px;
    }}
    .reason {{ margin-top: 8px; color: var(--muted); }}
    .cta-title {{
      margin-top: 12px;
      padding-top: 10px;
      border-top: 2px solid var(--line);
      color: var(--ink);
      font-size: 13px;
      font-weight: 700;
    }}
    .cta-list {{
      margin: 8px 0 0;
      padding-left: 18px;
    }}
    .cta-list li {{ margin: 7px 0; }}
    .cta-list span {{
      display: block;
      color: var(--muted);
      font-size: 12px;
    }}
    .guidance-panel {{
      display: grid;
      grid-template-columns: minmax(0, 1fr) 190px;
      gap: 24px;
      align-items: center;
      padding: 24px;
      border: 2px solid var(--line);
      border-radius: 8px;
      background: #f3f3f3;
      box-shadow: var(--shadow);
    }}
    .guidance-panel ul,
    .guidance-panel ol {{
      margin-top: 8px;
      padding-left: 22px;
    }}
    .subhead {{
      margin-top: 20px;
    }}
    .guidance-mascot {{
      width: 100%;
      height: 210px;
      object-fit: cover;
      object-position: center center;
      border: 2px solid var(--line);
      border-radius: 8px;
      background: #061923;
      box-shadow: 0 3px 0 #191a23;
    }}
    .notice {{
      border: 2px solid var(--line);
      border-radius: 8px;
      padding: 20px;
      background: var(--panel);
      box-shadow: 0 4px 0 #191a23;
    }}
    .check-panel {{
      display: grid;
      gap: 12px;
      padding: 24px;
      border: 2px solid var(--line);
      border-radius: 8px;
      background: #fff;
      box-shadow: var(--shadow);
    }}
    .check-panel p {{
      max-width: 920px;
      color: #343640;
    }}
    .error {{ border-color: #e3b5b5; color: var(--danger); background: #fff7f7; }}
    .final-cta {{
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 16px;
      margin-top: 10px;
      padding: 28px;
      border: 2px solid var(--line);
      border-radius: 8px;
      background: var(--accent);
      box-shadow: var(--shadow);
    }}
    .final-cta h2 {{
      margin: 0;
      color: var(--ink);
    }}
    .final-cta p {{
      max-width: 620px;
      margin-top: 8px;
      color: #343640;
    }}
    .cta-button {{
      display: inline-flex;
      align-items: center;
      justify-content: center;
      min-height: 54px;
      min-width: 190px;
      padding: 0 22px;
      color: #fff;
      background: var(--ink);
      font-weight: 700;
      text-decoration: none;
      border: 2px solid var(--line);
      border-radius: 8px;
    }}
    .cta-button:focus,
    .cta-button:hover {{
      background: #2f3140;
    }}
    @media (max-width: 760px) {{
      header {{ padding: 16px; }}
      main {{ padding: 16px; }}
      .hero {{ grid-template-columns: 1fr; padding: 22px; }}
      .intro-panel {{ grid-template-columns: 1fr; padding: 0 16px; }}
      .hero-art {{ min-height: 250px; }}
      .hero-art img {{ height: 250px; }}
      .section-head {{ min-height: 86px; padding: 14px 16px; }}
      .section-mascot {{ width: 108px; height: 76px; }}
      .guidance-panel {{ grid-template-columns: 1fr; }}
      .guidance-mascot {{ height: 190px; }}
      form, .summary, .grid, .survey-form, .choice-grid {{ grid-template-columns: 1fr; }}
      .survey-link {{ align-items: stretch; flex-direction: column; }}
      .final-cta {{ align-items: stretch; flex-direction: column; }}
      .cta-button {{ width: 100%; }}
      h1 {{ font-size: 32px; }}
    }}
  </style>
</head>
<body>
    <header>
      <div class="hero">
        <div>
          <span class="eyebrow">AI Career Counseling Report</span>
          <h1>FuturesImpactLAB 진로상담 리포트</h1>
          <p>성찰일지와 학과 정보, 다음 학기 개설 예정 교과목을 바탕으로 앞으로의 학습 방향과 진로 탐색을 함께 정리해보는 AI 기반 진로상담 보조 자료입니다.</p>
        </div>
        <div class="hero-art" aria-hidden="true">
          <img src="/new03.png" alt="">
        </div>
      </div>
      <div class="intro-panel">
        <section class="intro-card">
          <span class="section-kicker">Guide</span>
          <h2>진로는 한 번에 정답을 찾는 일이 아닙니다.</h2>
          <p>내가 어떤 분야에 관심이 있는지, 어떤 역량을 더 키우고 싶은지, 다음 학기에 어떤 선택을 해보면 좋을지 하나씩 확인해 가는 과정입니다.</p>
          <p>이 리포트는 그 과정을 조금 더 구체적으로 돕기 위해 성찰일지에 나타난 관심과 학과 배경, 교육과정과의 연결성을 함께 살펴봅니다.</p>
        </section>
        <section class="intro-card accent">
          <span class="label">이 리포트는 어떻게 만들어졌나요?</span>
          <p>입력한 익명 학생번호를 기준으로 성찰일지 응답, 소속·계열·학과 정보, 2026학년도 2학기 개설 예정 교과목, 관련 교육과정 문서를 함께 확인합니다.</p>
        </section>
      </div>
    </header>
    <main>
      <form method="post">
        <input name="student_id" value="{html.escape(student_id)}" placeholder="학번 입력" autocomplete="off" required>
        <button type="submit">리포트 생성</button>
      </form>
      <section class="survey-link">
        <div>
          <span class="label">성찰일지 데이터가 없나요?</span>
          <p>간단 문항에 답하면 응답 내용을 바탕으로 교과목/교육과정 추천 리포트를 만들 수 있습니다.</p>
        </div>
        <a class="cta-button" href="/survey">간단 응답 시작</a>
      </section>
      <section class="intro-card">
        <span class="section-kicker">Report Contents</span>
        <h2>리포트에서 확인할 수 있는 내용</h2>
        <ul class="intro-list">
          <li>성찰일지에 나타난 관심 분야와 향후 성장 방향을 정리한 진로목표 요약</li>
          <li>다음 학기에 실제 개설 예정인 과목 중 관심 분야와 연결되는 추천 교과목</li>
          <li>마이크로디그리와 연계 교육과정처럼 구조화된 학습 경로</li>
          <li>수업, 프로젝트, 비교과 활동, 포트폴리오 준비와 연결되는 진로 탐색 조언과 성찰 질문</li>
        </ul>
      </section>
      {body or "<section class='notice'>학번을 입력하면 학생/상담자 검토용 리포트가 여기에 표시됩니다.</section>"}
    </main>
</body>
</html>"""


class CareerReportHandler(BaseHTTPRequestHandler):
    def do_GET(self) -> None:
        parsed = urllib.parse.urlparse(self.path)
        static_path = parsed.path
        if static_path in IMAGE_FILES:
            self.respond_file(IMAGE_FILES[static_path], "image/png")
            return
        if static_path == "/admin":
            self.respond(render_admin_html())
            return
        if static_path == "/survey":
            params = urllib.parse.parse_qs(parsed.query)
            student_id = params.get("student_id", [""])[0].strip()
            self.respond(render_html(student_id=student_id, survey=True))
            return
        if static_path == "/go-course":
            self.handle_course_redirect(parsed.query)
            return
        self.respond(render_html())

    def do_POST(self) -> None:
        length = int(self.headers.get("Content-Length", "0"))
        body = self.rfile.read(length).decode("utf-8")
        params = urllib.parse.parse_qs(body)
        student_id = params.get("student_id", [""])[0].strip()
        if params.get("mode", [""])[0] == "survey":
            try:
                report = build_survey_report(params)
                write_analytics_event("survey_lookup", report["student"])
                self.respond(render_html(report=report, student_id=student_id))
            except Exception as exc:
                self.respond(render_html(error=str(exc), student_id=student_id, survey=True), status=400)
            return
        try:
            report = build_report(student_id)
            write_analytics_event("lookup", report["student"])
            self.respond(render_html(report=report, student_id=student_id))
        except KeyError as exc:
            self.respond(render_html(error=str(exc.args[0] if exc.args else exc), student_id=student_id, survey=True))
        except Exception as exc:
            self.respond(render_html(error=str(exc), student_id=student_id), status=400)

    def handle_course_redirect(self, query: str) -> None:
        params = urllib.parse.parse_qs(query)
        student_id = params.get("student_id", [""])[0]
        try:
            records, _headers = load_student_records()
            student = records.get(normalize_student_id(student_id))
            write_analytics_event("course_click", student, {"destination": COURSE_REGISTRATION_URL})
        except Exception:
            write_analytics_event("course_click", None, {"masked_student_id": "미정", "destination": COURSE_REGISTRATION_URL})
        self.send_response(302)
        self.send_header("Location", COURSE_REGISTRATION_URL)
        self.end_headers()

    def log_message(self, format: str, *args: Any) -> None:
        return

    def respond(self, content: str, status: int = 200) -> None:
        payload = content.encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Content-Length", str(len(payload)))
        self.end_headers()
        self.wfile.write(payload)

    def respond_file(self, path: Path, content_type: str) -> None:
        if not path.exists():
            self.send_error(404)
            return
        payload = path.read_bytes()
        self.send_response(200)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(payload)))
        self.end_headers()
        self.wfile.write(payload)


def run_server(host: str, port: int) -> None:
    server = ThreadingHTTPServer((host, port), CareerReportHandler)
    print(f"http://{host}:{port}")
    server.serve_forever()


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="학번 기반 진로상담 리포트 생성기")
    parser.add_argument("--student-id", help="리포트를 생성할 익명화 학번")
    parser.add_argument("--list-sample-ids", action="store_true", help="검증용으로 앞쪽 학번 5개를 마스킹해서 표시")
    parser.add_argument("--serve", action="store_true", help="웹앱 실행")
    parser.add_argument("--host", default="127.0.0.1")
    parser.add_argument("--port", type=int, default=8765)
    args = parser.parse_args(argv)

    if args.list_sample_ids:
        records, _headers = load_student_records()
        for sid in list(records.keys())[:5]:
            print(mask_student_id(sid))
        return 0

    if args.student_id:
        report = build_report(args.student_id)
        print(render_markdown(report))
        return 0

    run_server(args.host, args.port)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
